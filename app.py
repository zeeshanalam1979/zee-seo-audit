"""
SEO Site Auditor — Streamlit Web App
Built by Zeeshan Alam | SEO Specialist (22+ years)
Powered by Python + Claude AI
"""

import io
import re
import json
import time
import threading
from datetime import datetime
from urllib.parse import urljoin, urlparse

import requests
import streamlit as st
import pandas as pd
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, HRFlowable, PageBreak
)

# ─────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="SEO Site Auditor",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
#  CUSTOM CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
  /* Global */
  [data-testid="stAppViewContainer"] { background: #0f172a; }
  [data-testid="stSidebar"] { background: #1e293b; }
  h1, h2, h3, h4 { color: #e2e8f0 !important; }
  p, label, div { color: #cbd5e1; }

  /* Header banner */
  .hero {
    background: linear-gradient(135deg, #1e40af, #7c3aed);
    border-radius: 16px;
    padding: 2rem 2.5rem;
    margin-bottom: 1.5rem;
  }
  .hero h1 { font-size: 2rem; font-weight: 800; color: #fff !important; margin: 0; }
  .hero p  { color: rgba(255,255,255,0.75); margin: .4rem 0 0; font-size: .95rem; }

  /* Metric cards */
  .metric-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(150px, 1fr));
    gap: .85rem;
    margin: 1.2rem 0;
  }
  .metric-card {
    background: #1e293b;
    border: 1px solid #334155;
    border-radius: 12px;
    padding: 1.1rem;
    text-align: center;
  }
  .metric-card.alert { border-color: #ef4444; }
  .metric-card .val  { font-size: 1.9rem; font-weight: 700; color: #e2e8f0; }
  .metric-card .lbl  { font-size: .72rem; color: #94a3b8; margin-top: .25rem; }

  /* Score badge */
  .badge {
    display: inline-block;
    padding: .2rem .65rem;
    border-radius: 999px;
    font-weight: 700;
    font-size: .82rem;
    color: #fff;
  }

  /* Pills */
  .pill {
    display: inline-block;
    font-size: .7rem;
    padding: .15rem .45rem;
    border-radius: 999px;
    margin: .1rem;
    white-space: nowrap;
  }
  .pill-issue { background:#ef444420; color:#fca5a5; border:1px solid #ef4444; }
  .pill-warn  { background:#f59e0b20; color:#fcd34d; border:1px solid #f59e0b; }
  .pill-ok    { background:#22c55e20; color:#86efac; border:1px solid #22c55e; }

  /* Progress bar track */
  .stProgress > div > div { background: #334155; border-radius: 999px; }
  .stProgress > div > div > div { background: linear-gradient(90deg,#1e40af,#7c3aed); border-radius: 999px; }

  /* Buttons */
  .stButton > button {
    background: linear-gradient(135deg,#1e40af,#7c3aed) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    padding: .6rem 1.5rem !important;
  }
  .stButton > button:hover { opacity: .9; }

  /* Download button */
  .stDownloadButton > button {
    background: #1e293b !important;
    color: #60a5fa !important;
    border: 1px solid #334155 !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
  }

  /* Dataframe */
  [data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; }

  /* Sidebar */
  .sidebar-brand {
    background: linear-gradient(135deg,#1e40af,#7c3aed);
    border-radius: 12px;
    padding: 1rem;
    text-align: center;
    margin-bottom: 1rem;
  }
  .sidebar-brand h3 { color: #fff !important; font-size: 1rem; margin: 0; }
  .sidebar-brand p  { color: rgba(255,255,255,.7); font-size: .75rem; margin: .2rem 0 0; }

  /* Section divider */
  .section-title {
    font-size: .8rem;
    font-weight: 700;
    letter-spacing: .08em;
    text-transform: uppercase;
    color: #94a3b8;
    margin: 1.5rem 0 .75rem;
    padding-bottom: .4rem;
    border-bottom: 1px solid #334155;
  }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
TIMEOUT    = 10
USER_AGENT = "Mozilla/5.0 (compatible; SEO-Auditor/1.0)"
HEADERS    = {"User-Agent": USER_AGENT}

TITLE_MIN, TITLE_MAX = 30, 60
META_MIN,  META_MAX  = 120, 160
WORD_COUNT_MIN       = 300

SCORE_TITLE  = 20
SCORE_META   = 15
SCORE_H1     = 15
SCORE_IMAGES = 15
SCORE_WORDS  = 10
SCORE_CANON  = 10
SCORE_SCHEMA = 5


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def clean_url(url):
    p = urlparse(url)
    return p.scheme + "://" + p.netloc + p.path.rstrip("/")

def same_domain(url, base):
    return urlparse(url).netloc == urlparse(base).netloc

def grade(score):
    if score >= 90: return "A"
    if score >= 75: return "B"
    if score >= 60: return "C"
    if score >= 45: return "D"
    return "F"

def grade_color(score):
    if score >= 90: return "#22c55e"
    if score >= 75: return "#84cc16"
    if score >= 60: return "#f59e0b"
    if score >= 45: return "#f97316"
    return "#ef4444"


# ─────────────────────────────────────────────
#  PAGE ANALYSER
# ─────────────────────────────────────────────
def analyse_page(url, session):
    result = {
        "url": url, "status_code": None, "load_time_ms": None,
        "title": "", "title_len": 0, "meta_desc": "", "meta_desc_len": 0,
        "h1_count": 0, "h1_text": "", "h2_count": 0,
        "canonical": "", "robots_meta": "", "schema_types": [],
        "images_total": 0, "images_missing_alt": 0,
        "internal_links": 0, "external_links": 0,
        "word_count": 0, "issues": [], "warnings": [],
        "score": 0, "outbound_urls": [], "error": None,
    }
    try:
        t0   = time.time()
        resp = session.get(url, timeout=TIMEOUT, allow_redirects=True)
        result["load_time_ms"]  = round((time.time() - t0) * 1000)
        result["status_code"]   = resp.status_code

        if resp.status_code != 200:
            result["issues"].append(f"HTTP {resp.status_code}")
            return result

        soup = BeautifulSoup(resp.text, "html.parser")

        # Title
        t = soup.find("title")
        title = t.get_text(strip=True) if t else ""
        result["title"] = title;  result["title_len"] = len(title)

        # Meta desc
        m = soup.find("meta", attrs={"name": re.compile("^description$", re.I)})
        meta_desc = m.get("content", "").strip() if m else ""
        result["meta_desc"] = meta_desc;  result["meta_desc_len"] = len(meta_desc)

        # Headings
        h1s = soup.find_all("h1")
        result["h1_count"] = len(h1s)
        result["h1_text"]  = h1s[0].get_text(strip=True)[:120] if h1s else ""
        result["h2_count"] = len(soup.find_all("h2"))

        # Canonical / robots
        canon = soup.find("link", attrs={"rel": "canonical"})
        result["canonical"] = canon.get("href", "") if canon else ""
        robots = soup.find("meta", attrs={"name": re.compile("^robots$", re.I)})
        result["robots_meta"] = robots.get("content", "") if robots else ""

        # Schema
        schemas, schema_types = soup.find_all("script", attrs={"type": "application/ld+json"}), []
        for s in schemas:
            try:
                data = json.loads(s.string or "")
                tp   = data.get("@type", "")
                if tp: schema_types.append(tp if isinstance(tp, str) else ", ".join(tp))
            except Exception: pass
        result["schema_types"] = schema_types

        # Images
        imgs = soup.find_all("img")
        result["images_total"]       = len(imgs)
        result["images_missing_alt"] = sum(1 for i in imgs if not i.get("alt", "").strip())

        # Links
        base_domain   = urlparse(url).netloc
        internal, external, outbound = 0, 0, []
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if href.startswith(("#", "mailto:", "tel:")): continue
            abs_url = urljoin(url, href)
            if urlparse(abs_url).netloc == base_domain:
                internal += 1;  outbound.append(clean_url(abs_url))
            else:
                external += 1
        result["internal_links"] = internal
        result["external_links"] = external
        result["outbound_urls"]  = list(set(outbound))

        # Word count
        body = soup.find("body")
        if body:
            result["word_count"] = len(body.get_text(separator=" ", strip=True).split())

        # ── Scoring ──
        score = 100
        if not title:
            result["issues"].append("Missing title tag");  score -= SCORE_TITLE
        elif result["title_len"] < TITLE_MIN:
            result["warnings"].append(f"Title too short ({result['title_len']} chars)");  score -= SCORE_TITLE // 2
        elif result["title_len"] > TITLE_MAX:
            result["warnings"].append(f"Title too long ({result['title_len']} chars)");   score -= SCORE_TITLE // 2

        if not meta_desc:
            result["issues"].append("Missing meta description");  score -= SCORE_META
        elif result["meta_desc_len"] < META_MIN:
            result["warnings"].append(f"Meta too short ({result['meta_desc_len']} chars)");  score -= SCORE_META // 2
        elif result["meta_desc_len"] > META_MAX:
            result["warnings"].append(f"Meta too long ({result['meta_desc_len']} chars)");   score -= SCORE_META // 2

        if result["h1_count"] == 0:
            result["issues"].append("Missing H1 tag");  score -= SCORE_H1
        elif result["h1_count"] > 1:
            result["warnings"].append(f"Multiple H1 tags ({result['h1_count']})");  score -= SCORE_H1 // 2

        if result["images_missing_alt"] > 0:
            result["issues"].append(f"{result['images_missing_alt']} image(s) missing alt text")
            score -= min(SCORE_IMAGES, result["images_missing_alt"] * 3)

        if result["word_count"] < WORD_COUNT_MIN:
            result["warnings"].append(f"Low word count ({result['word_count']} words)")
            score -= SCORE_WORDS

        if not result["canonical"]:
            result["warnings"].append("No canonical tag");  score -= SCORE_CANON // 2

        if not result["schema_types"]:
            result["warnings"].append("No structured data (Schema.org)");  score -= SCORE_SCHEMA

        if result["load_time_ms"] and result["load_time_ms"] > 3000:
            result["issues"].append(f"Slow page: {result['load_time_ms']}ms")
        elif result["load_time_ms"] and result["load_time_ms"] > 1500:
            result["warnings"].append(f"Page could be faster: {result['load_time_ms']}ms")

        result["score"] = max(0, score)

    except requests.exceptions.Timeout:
        result["error"] = "Timeout";  result["issues"].append("Request timed out")
    except Exception as e:
        result["error"] = str(e);     result["issues"].append(f"Error: {e}")
    return result


# ─────────────────────────────────────────────
#  CRAWLER
# ─────────────────────────────────────────────
def crawl(start_url, max_pages, delay, progress_bar, status_text, log_container):
    session = requests.Session()
    session.headers.update(HEADERS)

    visited, queue, results = set(), [clean_url(start_url)], []
    log_lines = []

    while queue and len(visited) < max_pages:
        url = queue.pop(0)
        if url in visited: continue
        visited.add(url)

        idx  = len(visited)
        pct  = idx / max_pages
        progress_bar.progress(min(pct, 1.0))
        status_text.markdown(f"**Crawling [{idx}/{max_pages}]:** `{url}`")

        log_lines.append(f"✅ [{idx}] {url}")
        log_container.code("\n".join(log_lines[-15:]), language=None)

        page = analyse_page(url, session)
        results.append(page)

        for new_url in page.get("outbound_urls", []):
            if new_url not in visited and new_url not in queue:
                if same_domain(new_url, start_url):
                    queue.append(new_url)

        time.sleep(delay)

    progress_bar.progress(1.0)
    status_text.markdown(f"✅ **Crawl complete — {len(results)} pages audited**")
    return results


# ─────────────────────────────────────────────
#  SUMMARY
# ─────────────────────────────────────────────
def compute_summary(results, domain):
    total = len(results)
    if not total: return {}
    scores = [r["score"] for r in results]
    avg    = round(sum(scores) / total)
    return {
        "domain":        domain,
        "total":         total,
        "avg_score":     avg,
        "avg_grade":     grade(avg),
        "avg_load_ms":   round(sum(r["load_time_ms"] or 0 for r in results) / total),
        "broken_pages":  sum(1 for r in results if r["status_code"] and r["status_code"] >= 400),
        "missing_title": sum(1 for r in results if not r["title"]),
        "missing_meta":  sum(1 for r in results if not r["meta_desc"]),
        "missing_h1":    sum(1 for r in results if r["h1_count"] == 0),
        "multi_h1":      sum(1 for r in results if r["h1_count"] > 1),
        "missing_alt":   sum(1 for r in results if r["images_missing_alt"] > 0),
        "slow_pages":    sum(1 for r in results if r["load_time_ms"] and r["load_time_ms"] > 3000),
        "no_schema":     sum(1 for r in results if not r["schema_types"]),
        "no_canonical":  sum(1 for r in results if not r["canonical"]),
        "low_words":     sum(1 for r in results if r["word_count"] < WORD_COUNT_MIN),
    }


# ─────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────
def build_excel(results, summary):
    wb  = openpyxl.Workbook()
    thin = Side(style="thin", color="CBD5E1")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", fgColor="1E3A5F")
    hfont = Font(color="FFFFFF", bold=True)

    # Summary sheet
    ws = wb.active;  ws.title = "Summary"
    ws.append(["SEO Audit Summary", ""])
    ws.append(["Domain", summary.get("domain", "")])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append([])
    ws.append(["Metric", "Value"])
    for k, v in [
        ("Total Pages",          summary["total"]),
        ("Avg SEO Score",        summary["avg_score"]),
        ("Avg Grade",            summary["avg_grade"]),
        ("Avg Load Time (ms)",   summary["avg_load_ms"]),
        ("Broken Pages",         summary["broken_pages"]),
        ("Missing Titles",       summary["missing_title"]),
        ("Missing Meta Desc",    summary["missing_meta"]),
        ("Missing H1",           summary["missing_h1"]),
        ("Multiple H1",          summary["multi_h1"]),
        ("Images No Alt",        summary["missing_alt"]),
        ("Slow Pages (>3s)",     summary["slow_pages"]),
        ("No Schema Markup",     summary["no_schema"]),
        ("No Canonical",         summary["no_canonical"]),
        ("Low Word Count (<300)",summary["low_words"]),
    ]:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws["A5"].fill = hfill;  ws["A5"].font = hfont
    ws["B5"].fill = hfill;  ws["B5"].font = hfont

    # Pages sheet
    ws2 = wb.create_sheet("Pages")
    headers = ["URL","Status","Score","Grade","Load (ms)",
               "Title","Title Len","Meta Desc","Meta Len",
               "H1 Count","H1 Text","H2 Count","Canonical",
               "Robots Meta","Schema Types","Images Total",
               "Images Missing Alt","Internal Links","External Links",
               "Word Count","Issues","Warnings"]
    ws2.append(headers)
    for c in ws2[1]:
        c.fill = hfill;  c.font = hfont
        c.alignment = Alignment(horizontal="center");  c.border = bdr

    def sfill(s):
        if s >= 90: return PatternFill("solid", fgColor="BBF7D0")
        if s >= 75: return PatternFill("solid", fgColor="D9F99D")
        if s >= 60: return PatternFill("solid", fgColor="FEF08A")
        if s >= 45: return PatternFill("solid", fgColor="FED7AA")
        return PatternFill("solid", fgColor="FECACA")

    for r in results:
        ws2.append([
            r["url"], r["status_code"] or "", r["score"], grade(r["score"]),
            r["load_time_ms"] or "", r["title"], r["title_len"],
            r["meta_desc"], r["meta_desc_len"], r["h1_count"], r["h1_text"],
            r["h2_count"], r["canonical"], r["robots_meta"],
            ", ".join(r["schema_types"]), r["images_total"],
            r["images_missing_alt"], r["internal_links"], r["external_links"],
            r["word_count"], " | ".join(r["issues"]), " | ".join(r["warnings"]),
        ])
        sc = ws2.cell(row=ws2.max_row, column=3)
        sc.fill = sfill(r["score"]);  sc.font = Font(bold=True)

    for i, w in enumerate([50,8,8,7,10,40,10,60,10,8,40,8,40,20,30,12,18,14,14,10,60,60], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(headers)):
        for c in row:
            c.border = bdr;  c.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.freeze_panes = "A2";  ws2.auto_filter.ref = ws2.dimensions

    buf = io.BytesIO();  wb.save(buf);  buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  PDF EXPORT
# ─────────────────────────────────────────────
def build_pdf(results, summary):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    t_style = ParagraphStyle("T", parent=styles["Title"], fontSize=20,
                              textColor=colors.HexColor("#1e40af"), spaceAfter=6)
    s_style = ParagraphStyle("S", parent=styles["Normal"], fontSize=9,
                              textColor=colors.grey, spaceAfter=14)
    h_style = ParagraphStyle("H", parent=styles["Heading2"], fontSize=12,
                              textColor=colors.HexColor("#1e40af"), spaceBefore=16, spaceAfter=8)
    c_style = ParagraphStyle("C", parent=styles["Normal"], fontSize=7, leading=9)

    def sc(s):
        if s >= 90: return colors.HexColor("#22c55e")
        if s >= 75: return colors.HexColor("#84cc16")
        if s >= 60: return colors.HexColor("#f59e0b")
        if s >= 45: return colors.HexColor("#f97316")
        return colors.HexColor("#ef4444")

    story.append(Spacer(1, .5*cm))
    story.append(Paragraph("SEO Audit Report", t_style))
    story.append(Paragraph(
        f"Domain: <b>{summary.get('domain','')}</b>  |  "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"{summary['total']} pages audited", s_style))
    story.append(HRFlowable(width="100%", color=colors.HexColor("#e2e8f0")))
    story.append(Paragraph("Executive Summary", h_style))

    card_data = [
        ["Total Pages", summary["total"], "Avg SEO Score", f"{summary['avg_score']} ({summary['avg_grade']})"],
        ["Avg Load Time", f"{summary['avg_load_ms']} ms", "Broken Pages", summary["broken_pages"]],
        ["Missing Titles", summary["missing_title"], "Missing Meta", summary["missing_meta"]],
        ["Missing H1", summary["missing_h1"], "Images No Alt", summary["missing_alt"]],
        ["Slow Pages (>3s)", summary["slow_pages"], "No Schema", summary["no_schema"]],
    ]
    ts = TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#f8fafc")),
        ("FONTNAME",  (0,0),(-1,-1),"Helvetica"),
        ("FONTSIZE",  (0,0),(-1,-1),9),
        ("FONTNAME",  (1,0),(1,-1),"Helvetica-Bold"),
        ("FONTNAME",  (3,0),(3,-1),"Helvetica-Bold"),
        ("FONTSIZE",  (1,0),(1,-1),11),
        ("FONTSIZE",  (3,0),(3,-1),11),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.HexColor("#f1f5f9"),colors.HexColor("#e2e8f0")]),
        ("BOX",       (0,0),(-1,-1),.5,colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0,0),(-1,-1),.25,colors.HexColor("#cbd5e1")),
        ("ALIGN",     (1,0),(1,-1),"CENTER"),
        ("ALIGN",     (3,0),(3,-1),"CENTER"),
        ("VALIGN",    (0,0),(-1,-1),"MIDDLE"),
        ("PADDING",   (0,0),(-1,-1),8),
    ])
    story.append(Table(card_data, colWidths=[4.2*cm,2.5*cm,4.2*cm,2.5*cm], style=ts))

    story.append(PageBreak())
    story.append(Paragraph("Page-by-Page Results", h_style))
    hdr = ["URL","Status","Score","Load\n(ms)","Title\nLen","Meta\nLen","H1s","Alt\nMiss","Words","Top Issues"]
    rows = [hdr]
    for r in results:
        ti = (r["issues"] + r["warnings"])[:2]
        rows.append([
            Paragraph((r["url"][:47]+"…") if len(r["url"])>50 else r["url"], c_style),
            r["status_code"] or "–", r["score"], r["load_time_ms"] or "–",
            r["title_len"], r["meta_desc_len"], r["h1_count"],
            r["images_missing_alt"], r["word_count"],
            Paragraph(("; ".join(ti)[:80]) if ti else "✓ OK", c_style),
        ])
    pts = TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0,0),(-1,0),colors.white),
        ("FONTNAME",  (0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",  (0,0),(-1,-1),7),
        ("ALIGN",     (1,0),(-2,-1),"CENTER"),
        ("VALIGN",    (0,0),(-1,-1),"TOP"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f8fafc")]),
        ("BOX",       (0,0),(-1,-1),.5,colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0,0),(-1,-1),.25,colors.HexColor("#e2e8f0")),
        ("PADDING",   (0,0),(-1,-1),4),
    ])
    for i, r in enumerate(results, 1):
        pts.add("TEXTCOLOR",(2,i),(2,i),sc(r["score"]))
        pts.add("FONTNAME", (2,i),(2,i),"Helvetica-Bold")
    story.append(Table(rows, colWidths=[5*cm,1.2*cm,1.2*cm,1.3*cm,1.2*cm,1.2*cm,0.8*cm,1.3*cm,1.2*cm,4.8*cm],
                        repeatRows=1, style=pts))
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  HTML EXPORT
# ─────────────────────────────────────────────
def build_html(results, summary, domain):
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    def badge(s):
        return f'<span class="badge" style="background:{grade_color(s)}">{s} ({grade(s)})</span>'

    def pills(issues, warnings):
        out = "".join(f'<span class="pill pill-issue">{i}</span>' for i in issues)
        out += "".join(f'<span class="pill pill-warn">{w}</span>' for w in warnings)
        return out or '<span class="pill pill-ok">✓ No issues</span>'

    rows = ""
    for r in results:
        rows += f"""<tr
          data-status="{r['status_code'] or 0}" data-score="{r['score']}"
          data-load="{r['load_time_ms'] or 0}" data-titlelen="{r['title_len']}"
          data-metalen="{r['meta_desc_len']}" data-h1="{r['h1_count']}"
          data-altmissing="{r['images_missing_alt']}" data-words="{r['word_count']}"
          data-issues="{len(r['issues'])+len(r['warnings'])}">
          <td><a href="{r['url']}" target="_blank">{r['url']}</a></td>
          <td class="c">{r['status_code'] or '–'}</td>
          <td class="c">{badge(r['score'])}</td>
          <td class="c">{r['load_time_ms'] or '–'} ms</td>
          <td class="c">{r['title_len']}</td>
          <td class="c">{r['meta_desc_len']}</td>
          <td class="c">{r['h1_count']}</td>
          <td class="c">{r['images_missing_alt']}/{r['images_total']}</td>
          <td class="c">{r['word_count']}</td>
          <td>{pills(r['issues'], r['warnings'])}</td>
        </tr>"""

    def card(val, lbl, alert=False):
        cls = "card alert" if alert else "card"
        return f'<div class="{cls}"><div class="val">{val}</div><div class="lbl">{lbl}</div></div>'

    cards = (
        card(summary["total"], "Pages Audited") +
        card(f"{summary['avg_score']} ({summary['avg_grade']})", "Avg SEO Score") +
        card(f"{summary['avg_load_ms']} ms", "Avg Load Time") +
        card(summary["broken_pages"],  "Broken Pages",    summary["broken_pages"] > 0) +
        card(summary["missing_title"], "Missing Titles",  summary["missing_title"] > 0) +
        card(summary["missing_meta"],  "Missing Meta",    summary["missing_meta"] > 0) +
        card(summary["missing_h1"],    "Missing H1",      summary["missing_h1"] > 0) +
        card(summary["missing_alt"],   "Images No Alt",   summary["missing_alt"] > 0) +
        card(summary["slow_pages"],    "Slow Pages (>3s)") +
        card(summary["no_schema"],     "No Schema Markup")
    )

    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>SEO Audit – {domain}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0f172a;color:#e2e8f0}}
header{{background:linear-gradient(135deg,#1e40af,#7c3aed);padding:2rem}}
header h1{{font-size:1.8rem;font-weight:800}}
header p{{opacity:.75;font-size:.9rem;margin-top:.3rem}}
.wrap{{max-width:1400px;margin:0 auto;padding:2rem}}
.sec{{font-size:.75rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;color:#94a3b8;margin:1.5rem 0 .75rem;padding-bottom:.4rem;border-bottom:1px solid #334155}}
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(145px,1fr));gap:.85rem;margin-bottom:2rem}}
.card{{background:#1e293b;border:1px solid #334155;border-radius:12px;padding:1.1rem;text-align:center}}
.card.alert{{border-color:#ef4444}}
.card .val{{font-size:1.8rem;font-weight:700}}
.card .lbl{{font-size:.72rem;color:#94a3b8;margin-top:.25rem}}
.tbl-wrap{{overflow-x:auto;border-radius:12px;border:1px solid #334155}}
table{{width:100%;border-collapse:collapse;font-size:.82rem}}
thead{{background:#1e293b}}
th{{padding:.75rem 1rem;text-align:left;color:#94a3b8;font-weight:600;white-space:nowrap}}
th.s{{cursor:pointer;user-select:none;transition:color .15s}}
th.s:hover{{color:#e2e8f0}}
th.s::after{{content:' ⇅';font-size:.7rem;opacity:.4}}
th.asc::after{{content:' ▲';opacity:1;color:#60a5fa}}
th.desc::after{{content:' ▼';opacity:1;color:#60a5fa}}
th.asc,th.desc{{color:#e2e8f0}}
tbody tr{{border-top:1px solid #1e293b}}
tbody tr:hover{{background:#1e293b88}}
td{{padding:.65rem 1rem;vertical-align:top}}
td a{{color:#60a5fa;text-decoration:none;word-break:break-all}}
td a:hover{{text-decoration:underline}}
.c{{text-align:center}}
.badge{{display:inline-block;padding:.2rem .6rem;border-radius:999px;font-weight:700;font-size:.8rem;color:#fff}}
.pill{{display:inline-block;font-size:.7rem;padding:.15rem .45rem;border-radius:999px;margin:.1rem;white-space:nowrap}}
.pill-issue{{background:#ef444420;color:#fca5a5;border:1px solid #ef4444}}
.pill-warn{{background:#f59e0b20;color:#fcd34d;border:1px solid #f59e0b}}
.pill-ok{{background:#22c55e20;color:#86efac;border:1px solid #22c55e}}
.hint{{font-size:.75rem;color:#475569;margin-bottom:.6rem}}
footer{{text-align:center;padding:2rem;color:#475569;font-size:.8rem}}
</style></head><body>
<header>
  <h1>🔍 SEO Audit Report</h1>
  <p>{domain} · {now} · {summary['total']} pages audited · Built by Zeeshan Alam</p>
</header>
<div class="wrap">
  <div class="sec">Summary</div>
  <div class="grid">{cards}</div>
  <div class="sec">Page-by-Page Results</div>
  <p class="hint">Click any column header to sort ↑↓</p>
  <div class="tbl-wrap">
    <table id="t">
      <thead><tr>
        <th>URL</th>
        <th class="s" data-col="status">Status</th>
        <th class="s" data-col="score">Score</th>
        <th class="s" data-col="load">Load Time</th>
        <th class="s" data-col="titlelen">Title Len</th>
        <th class="s" data-col="metalen">Meta Len</th>
        <th class="s" data-col="h1">H1s</th>
        <th class="s" data-col="altmissing">Alt Missing</th>
        <th class="s" data-col="words">Words</th>
        <th class="s" data-col="issues">Issues</th>
      </tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</div>
<footer>SEO Auditor · Built by Zeeshan Alam · {now}</footer>
<script>
(function(){{
  var tbody=document.querySelector('#t tbody'),
      ths=document.querySelectorAll('th.s'),state={{}};
  ths.forEach(function(th){{
    th.addEventListener('click',function(){{
      var col=th.dataset.col,dir=state[col]==='desc'?'asc':'desc';
      state[col]=dir;
      ths.forEach(function(h){{h.classList.remove('asc','desc')}});
      th.classList.add(dir);
      var rows=Array.from(tbody.querySelectorAll('tr'));
      rows.sort(function(a,b){{
        var av=parseFloat(a.dataset[col])||0,bv=parseFloat(b.dataset[col])||0;
        return dir==='asc'?av-bv:bv-av;
      }});
      rows.forEach(function(r){{tbody.appendChild(r)}});
    }});
  }});
}})();
</script>
</body></html>"""


# ─────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div class="sidebar-brand">
      <h3>🔍 SEO Site Auditor</h3>
      <p>Built by Zeeshan Alam</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-title">Crawl Settings</div>', unsafe_allow_html=True)
    max_pages = st.slider("Max Pages to Crawl", 5, 100, 50, 5)
    delay     = st.slider("Delay Between Requests (s)", 0.2, 2.0, 0.5, 0.1)

    st.markdown('<div class="section-title">About</div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:.82rem;color:#94a3b8;line-height:1.6">
    ✅ Title & Meta tags<br>
    ✅ H1 / H2 structure<br>
    ✅ Image alt attributes<br>
    ✅ Page load speed<br>
    ✅ Schema markup<br>
    ✅ Canonical tags<br>
    ✅ Broken pages (4xx/5xx)<br>
    ✅ Word count / thin content<br>
    ✅ SEO Score + A–F Grade<br>
    ✅ HTML · Excel · PDF reports
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-title">Connect</div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:.82rem;color:#94a3b8">
    👤 <a href="https://linkedin.com/in/zeeshan-alam-seo-expert/" target="_blank"
         style="color:#60a5fa">Zeeshan Alam on LinkedIn</a><br>
    📧 zeeshan.alam1@gmail.com
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  MAIN UI
# ─────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <h1>🔍 SEO Site Auditor</h1>
  <p>Enter any website URL · Crawl up to 100 pages · Get instant HTML, Excel & PDF reports</p>
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns([4, 1])
with col1:
    url_input = st.text_input("", placeholder="https://example.com", label_visibility="collapsed")
with col2:
    run_btn = st.button("🚀 Start Audit", use_container_width=True)

# ─────────────────────────────────────────────
#  RUN AUDIT
# ─────────────────────────────────────────────
if run_btn:
    if not url_input.strip():
        st.error("Please enter a URL first.")
    else:
        start_url = url_input.strip()
        if not start_url.startswith("http"):
            start_url = "https://" + start_url

        domain = urlparse(start_url).netloc

        st.markdown('<div class="section-title">Live Crawl Progress</div>', unsafe_allow_html=True)
        progress_bar  = st.progress(0)
        status_text   = st.empty()
        log_container = st.empty()

        with st.spinner(""):
            results = crawl(start_url, max_pages, delay, progress_bar, status_text, log_container)

        summary = compute_summary(results, domain)

        # ── Metric Cards ──
        st.markdown('<div class="section-title">Audit Summary</div>', unsafe_allow_html=True)

        c1,c2,c3,c4,c5 = st.columns(5)
        c1.metric("Pages Audited",  summary["total"])
        c2.metric("Avg SEO Score",  f"{summary['avg_score']} ({summary['avg_grade']})")
        c3.metric("Avg Load Time",  f"{summary['avg_load_ms']} ms")
        c4.metric("Broken Pages",   summary["broken_pages"])
        c5.metric("Slow Pages",     summary["slow_pages"])

        c6,c7,c8,c9,c10 = st.columns(5)
        c6.metric("Missing Titles", summary["missing_title"])
        c7.metric("Missing Meta",   summary["missing_meta"])
        c8.metric("Missing H1",     summary["missing_h1"])
        c9.metric("Images No Alt",  summary["missing_alt"])
        c10.metric("No Schema",     summary["no_schema"])

        # ── DataFrame Table ──
        st.markdown('<div class="section-title">Page-by-Page Results</div>', unsafe_allow_html=True)

        df = pd.DataFrame([{
            "URL":          r["url"],
            "Status":       r["status_code"] or "–",
            "Score":        r["score"],
            "Grade":        grade(r["score"]),
            "Load (ms)":    r["load_time_ms"] or 0,
            "Title Len":    r["title_len"],
            "Meta Len":     r["meta_desc_len"],
            "H1s":          r["h1_count"],
            "Alt Missing":  r["images_missing_alt"],
            "Words":        r["word_count"],
            "Issues":       len(r["issues"]),
            "Warnings":     len(r["warnings"]),
        } for r in results])

        st.dataframe(
            df,
            use_container_width=True,
            height=420,
            column_config={
                "Score": st.column_config.ProgressColumn(
                    "Score", min_value=0, max_value=100, format="%d"
                ),
                "URL": st.column_config.LinkColumn("URL"),
            }
        )

        # ── Download Reports ──
        st.markdown('<div class="section-title">Download Reports</div>', unsafe_allow_html=True)
        slug = domain.replace(".", "_").replace("-", "_")

        d1, d2, d3 = st.columns(3)

        with d1:
            html_bytes = build_html(results, summary, domain).encode("utf-8")
            st.download_button(
                "📄 Download HTML Report",
                data=html_bytes,
                file_name=f"seo_report_{slug}.html",
                mime="text/html",
                use_container_width=True,
            )

        with d2:
            xlsx_bytes = build_excel(results, summary)
            st.download_button(
                "📊 Download Excel Report",
                data=xlsx_bytes,
                file_name=f"seo_report_{slug}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with d3:
            pdf_bytes = build_pdf(results, summary)
            st.download_button(
                "📋 Download PDF Report",
                data=pdf_bytes,
                file_name=f"seo_report_{slug}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        st.success(f"✅ Audit complete! {summary['total']} pages · Avg Score: {summary['avg_score']} ({summary['avg_grade']})")
